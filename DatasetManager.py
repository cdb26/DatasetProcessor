import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os, re, shutil, threading, datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False

# ─── Windows DPI + CTkToplevel icon fix ──────────────────────────────────────
import sys as _sys
if _sys.platform == "win32":
    try:
        import ctypes as _ct
        try:    _ct.windll.shcore.SetProcessDpiAwareness(2)
        except: _ct.windll.user32.SetProcessDPIAware()
    except: pass
    try:
        import customtkinter.windows.ctk_toplevel as _ctl
        _orig_ctl = _ctl.CTkToplevel.__init__
        def _safe_ctl(self, *a, **kw):
            _orig_ctl(self, *a, **kw)
            try:
                for _id in str(self.tk.call("after","info")).split():
                    try: self.after_cancel(_id)
                    except: pass
            except: pass
        _ctl.CTkToplevel.__init__ = _safe_ctl
    except: pass

# ─── Theme ────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
ACCENT="#4F8EF7"; ACCENT2="#7C5CFC"; BG_DARK="#0F1117"; BG_MID="#161B27"
BG_CARD="#1E2535"; BG_FIELD="#252D3D"; TEXT_MAIN="#E8EDF5"; TEXT_DIM="#7A8599"
SUCCESS="#3DD68C"; WARNING="#F5A623"; DANGER="#F75C5C"; BORDER="#2E3A50"

# ─── Filename helpers ─────────────────────────────────────────────────────────
FILENAME_RE = re.compile(
    r"^(\d{6})_([\d.]+m)_(\d)_(close|medium|far)_(dim|well)_(\d{4})(_depth)?(\.(?:jpg|png))$",
    re.IGNORECASE)

def parse_filename(name):
    m = FILENAME_RE.match(name)
    if not m: return None
    return {"room":m.group(1),"height":m.group(2),"angle":m.group(3),
            "distance":m.group(4),"lighting":m.group(5),"sequence":m.group(6),
            "is_depth":m.group(7) is not None,"ext":m.group(8).lower(),"original":name}

def build_filename(p):
    d = "_depth" if p.get("is_depth") else ""
    return f"{p['room']}_{p['height']}_{p['angle']}_{p['distance']}_{p['lighting']}_{p['sequence']}{d}{p['ext']}"

def group_key(p): return f"{p['room']}_{p['height']}_{p['angle']}_{p['distance']}_{p['lighting']}"
def base_key(p):  return f"{p['room']}_{p['height']}_{p['angle']}_{p['distance']}_{p['lighting']}_{p['sequence']}"

ANGLE_LABEL={"1":"Ortho","2":"Diagonal","3":"Top-down"}
LIGHT_LABEL={"dim":"Dim","well":"Well-lit"}
HEIGHT_OPTS=["0.8m","1.2m","1.6m"]; ANGLE_OPTS=["1","2","3"]
DIST_OPTS=["close","medium","far"]; LIGHT_OPTS=["dim","well"]

def walk_images(root):
    for dp,_,files in os.walk(root):
        for f in sorted(files):
            if f.lower().endswith((".jpg",".png")):
                yield os.path.join(dp,f), os.path.relpath(dp,root), f

def _excel_auto_name(matches):
    return (matches[0]["parts"]["room"]+".xlsx") if matches else "record.xlsx"

def _load_object_counts_from_xlsx(path):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb.active
    rows=list(ws.iter_rows(values_only=True))
    if len(rows)<2: return {},[]
    headers=[str(h).strip() if h else "" for h in rows[0]]
    try: kc=headers.index("Start Filename")
    except ValueError: return {},[]
    META={"Date","Floor","Room","Height (m)","Distance","Angle","Lighting","Resolution",
          "RGB Format","Depth Format","Start Filename","End Filename","# Images",
          "Est. Total Objects","Object Class","Notes","TOTAL"}
    obj_cols=[(i,headers[i]) for i in range(len(headers)) if headers[i] and headers[i] not in META]
    obj_names=[n for _,n in obj_cols]; counts={}
    for row in rows[1:]:
        fv=row[kc] if kc<len(row) else None
        if not fv: continue
        p=parse_filename(str(fv).strip())
        if not p: continue
        bk=base_key(p)
        counts[bk]={n:int(row[i] or 0) if i<len(row) else 0 for i,n in obj_cols}
    wb.close(); return counts,obj_names


# ═══════════════════════════════════════════════════════════════════════════════
class DatasetManagerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Dataset Manager"); self.geometry("1280x840")
        self.minsize(980,680); self.configure(fg_color=BG_DARK)
        self.dataset_path=tk.StringVar(); self.status_var=tk.StringVar(value="Ready")
        self._build_ui()

    def _build_ui(self):
        hdr=ctk.CTkFrame(self,fg_color=BG_MID,corner_radius=0,height=60)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="⬡  DATASET MANAGER",font=("Courier New",17,"bold"),
                     text_color=ACCENT).pack(side="left",padx=24)
        ctk.CTkLabel(hdr,text="Image filename toolkit for structured data collection",
                     font=("Courier New",11),text_color=TEXT_DIM).pack(side="left",padx=4)

        pr=ctk.CTkFrame(self,fg_color=BG_MID,corner_radius=0,height=52)
        pr.pack(fill="x"); pr.pack_propagate(False)
        ctk.CTkLabel(pr,text="Dataset root:",font=("Courier New",11),
                     text_color=TEXT_DIM).pack(side="left",padx=(20,6),pady=14)
        ctk.CTkEntry(pr,textvariable=self.dataset_path,font=("Courier New",11),
                     fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=BORDER,
                     width=560).pack(side="left",pady=14)
        ctk.CTkButton(pr,text="Browse",width=90,fg_color=ACCENT,hover_color=ACCENT2,
                      font=("Courier New",11,"bold"),command=self._browse).pack(side="left",padx=8)

        self.tabs=ctk.CTkTabview(self,fg_color=BG_CARD,
            segmented_button_fg_color=BG_MID,segmented_button_selected_color=ACCENT,
            segmented_button_unselected_color=BG_MID,
            segmented_button_selected_hover_color=ACCENT2,
            text_color=TEXT_MAIN,corner_radius=8)
        self.tabs.pack(fill="both",expand=True,padx=16,pady=(10,0))
        for n in ("✦  Rename","⊞  Filter / Search","⇄  Move by Room",
                  "⟳  Resequence","◈  Preview"):
            self.tabs.add(n)
        self._build_rename_tab(self.tabs.tab("✦  Rename"))
        self._build_filter_tab(self.tabs.tab("⊞  Filter / Search"))
        self._build_move_tab(self.tabs.tab("⇄  Move by Room"))
        self._build_resequence_tab(self.tabs.tab("⟳  Resequence"))
        self._build_preview_tab(self.tabs.tab("◈  Preview"))

        sb=ctk.CTkFrame(self,fg_color=BG_MID,corner_radius=0,height=30)
        sb.pack(fill="x",side="bottom"); sb.pack_propagate(False)
        ctk.CTkLabel(sb,textvariable=self.status_var,font=("Courier New",10),
                     text_color=TEXT_DIM).pack(side="left",padx=16)

    def _browse(self):
        d=filedialog.askdirectory(title="Select dataset root folder")
        if d: self.dataset_path.set(d); self._set_status(f"Dataset root: {d}")

    def _set_status(self,msg): self.status_var.set(msg)
    def _sl(self,p,t,px=0):
        ctk.CTkLabel(p,text=t,font=("Courier New",9,"bold"),text_color=ACCENT
                     ).pack(anchor="w",padx=px,pady=(8,2))
    def _log(self,w,t):
        w.configure(state="normal"); w.insert("end",t+"\n"); w.see("end"); w.configure(state="disabled")
    def _clr(self,w):
        w.configure(state="normal"); w.delete("1.0","end"); w.configure(state="disabled")
    def _get_root(self):
        r=self.dataset_path.get().strip()
        if not r or not os.path.isdir(r):
            messagebox.showerror("Error","Please select a valid dataset root folder."); return None
        return r

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 1 — RENAME
    # ══════════════════════════════════════════════════════════════════════════
    def _build_rename_tab(self,parent):
        parent.configure(fg_color=BG_CARD)
        L=ctk.CTkFrame(parent,fg_color=BG_CARD,corner_radius=0)
        L.pack(side="left",fill="both",expand=True,padx=(12,6),pady=12)
        R=ctk.CTkFrame(parent,fg_color=BG_CARD,corner_radius=0,width=280)
        R.pack(side="right",fill="y",padx=(0,12),pady=12); R.pack_propagate(False)

        self._sl(L,"CURRENT  →  NEW  FIELD  VALUES")
        g=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); g.pack(fill="x",pady=(6,0))
        self._rename_vars={}
        for i,(lbl,key,opts) in enumerate([
            ("Floor + Room  (FFRRRR)","new_room","e.g. 070202"),
            ("Height","new_height",HEIGHT_OPTS),("Angle","new_angle",ANGLE_OPTS),
            ("Distance","new_dist",DIST_OPTS),("Lighting","new_light",LIGHT_OPTS)]):
            ctk.CTkLabel(g,text=lbl,font=("Courier New",10),text_color=TEXT_DIM,
                         anchor="w").grid(row=i,column=0,sticky="w",padx=14,pady=6)
            var=tk.StringVar(); self._rename_vars[key]=var
            if isinstance(opts,list):
                mb=ctk.CTkOptionMenu(g,variable=var,values=["(keep)"]+opts,fg_color=BG_CARD,
                    button_color=ACCENT,button_hover_color=ACCENT2,text_color=TEXT_MAIN,
                    font=("Courier New",11),width=180)
                mb.set("(keep)"); mb.grid(row=i,column=1,sticky="w",padx=10,pady=6)
            else:
                ctk.CTkEntry(g,textvariable=var,font=("Courier New",11),fg_color=BG_CARD,
                    text_color=TEXT_MAIN,border_color=BORDER,placeholder_text=opts,
                    width=180).grid(row=i,column=1,sticky="w",padx=10,pady=6)

        so=ctk.CTkFrame(L,fg_color="transparent"); so.pack(fill="x",pady=(10,0))
        self._sl(so,"SEQUENCE")
        sc=ctk.CTkFrame(so,fg_color=BG_FIELD,corner_radius=8); sc.pack(fill="x")
        self._seq_mode=tk.StringVar(value="all")
        rr=ctk.CTkFrame(sc,fg_color="transparent"); rr.pack(anchor="w",padx=14,pady=(8,4))
        ctk.CTkLabel(rr,text="Apply to:",font=("Courier New",10),text_color=TEXT_DIM
                     ).pack(side="left",padx=(0,10))
        for val,txt,fg,hov in [("all","All sequences",ACCENT,ACCENT2),
                                ("selected","Selected range",ACCENT2,ACCENT)]:
            ctk.CTkRadioButton(rr,text=txt,variable=self._seq_mode,value=val,
                font=("Courier New",11),text_color=TEXT_MAIN,fg_color=fg,hover_color=hov,
                command=self._toggle_seq).pack(side="left",padx=(0,18))
        self._seq_rf=ctk.CTkFrame(sc,fg_color="transparent")
        self._seq_rf.pack(anchor="w",padx=14,pady=(0,10))
        self._seq_s=tk.StringVar(); self._seq_e=tk.StringVar()
        for lbl,var,ph in [("From",self._seq_s,"e.g. 0617"),("To  ",self._seq_e,"e.g. 0703")]:
            ctk.CTkLabel(self._seq_rf,text=lbl,font=("Courier New",10),
                         text_color=TEXT_DIM).pack(side="left",padx=(0,4))
            ctk.CTkEntry(self._seq_rf,textvariable=var,width=100,font=("Courier New",11),
                fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=BORDER,
                placeholder_text=ph).pack(side="left",padx=(0,14))
        self._toggle_seq()

        fr=ctk.CTkFrame(L,fg_color="transparent"); fr.pack(fill="x",pady=(10,0))
        self._sl(fr,"ONLY RENAME FILES MATCHING  (leave blank = all)")
        fi=ctk.CTkFrame(fr,fg_color=BG_FIELD,corner_radius=8); fi.pack(fill="x")
        self._rfilter_vars={}
        for j,(lbl,key,opts) in enumerate([
            ("Room","rf_room","e.g. 070701"),("Height","rf_height",HEIGHT_OPTS),
            ("Angle","rf_angle",ANGLE_OPTS),("Distance","rf_dist",DIST_OPTS),
            ("Lighting","rf_light",LIGHT_OPTS)]):
            ctk.CTkLabel(fi,text=lbl,font=("Courier New",10),text_color=TEXT_DIM
                         ).grid(row=j//3,column=(j%3)*2,sticky="w",padx=10,pady=6)
            v2=tk.StringVar(); self._rfilter_vars[key]=v2
            if isinstance(opts,list):
                mb2=ctk.CTkOptionMenu(fi,variable=v2,values=["(any)"]+opts,fg_color=BG_CARD,
                    button_color=ACCENT2,button_hover_color=ACCENT,text_color=TEXT_MAIN,
                    font=("Courier New",11),width=130)
                mb2.set("(any)"); mb2.grid(row=j//3,column=(j%3)*2+1,sticky="w",padx=6,pady=6)
            else:
                ctk.CTkEntry(fi,textvariable=v2,width=100,font=("Courier New",11),
                    fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=BORDER,
                    placeholder_text=opts).grid(row=j//3,column=(j%3)*2+1,sticky="w",padx=6,pady=6)

        self._sl(R,"OPTIONS")
        of=ctk.CTkFrame(R,fg_color=BG_FIELD,corner_radius=8); of.pack(fill="x",pady=(4,0))
        self._dry_run=tk.BooleanVar(value=True); self._backup=tk.BooleanVar(value=True)
        self._both_exts=tk.BooleanVar(value=True)
        for txt,var in [("Dry run (preview only)",self._dry_run),
                         ("Backup originals",self._backup),
                         ("Rename .jpg + .png pairs",self._both_exts)]:
            ctk.CTkCheckBox(of,text=txt,variable=var,font=("Courier New",11),
                text_color=TEXT_MAIN,fg_color=ACCENT,hover_color=ACCENT2
                ).pack(anchor="w",padx=14,pady=7)
        ctk.CTkButton(R,text="⟳  Preview Changes",height=38,fg_color=BG_FIELD,
            border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",12,"bold"),text_color=ACCENT,
            command=self._preview_rename).pack(fill="x",pady=(18,6))
        ctk.CTkButton(R,text="✔  Apply Rename",height=44,fg_color=ACCENT,hover_color=ACCENT2,
            font=("Courier New",13,"bold"),text_color="white",
            command=self._apply_rename).pack(fill="x",pady=6)

        self._sl(L,"OPERATION LOG")
        self._rename_log=ctk.CTkTextbox(L,height=160,font=("Courier New",10),
            fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=BORDER,corner_radius=6)
        self._rename_log.pack(fill="both",expand=True,pady=(4,0))

    def _toggle_seq(self):
        s="normal" if self._seq_mode.get()=="selected" else "disabled"
        for c in self._seq_rf.winfo_children(): c.configure(state=s)

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 2 — FILTER / SEARCH
    # ══════════════════════════════════════════════════════════════════════════
    def _build_filter_tab(self,parent):
        parent.configure(fg_color=BG_CARD)
        top=ctk.CTkFrame(parent,fg_color=BG_CARD); top.pack(fill="x",padx=12,pady=12)
        self._sl(top,"FILTER  CRITERIA  (leave field as '(any)' to skip)")
        cr=ctk.CTkFrame(top,fg_color=BG_FIELD,corner_radius=8); cr.pack(fill="x")
        self._filter_vars={}
        for idx,(lbl,key,opts) in enumerate([
            ("Room","f_room",None),("Height","f_height",HEIGHT_OPTS),
            ("Angle","f_angle",ANGLE_OPTS),("Distance","f_dist",DIST_OPTS),
            ("Lighting","f_light",LIGHT_OPTS)]):
            ctk.CTkLabel(cr,text=lbl,font=("Courier New",10),text_color=TEXT_DIM,
                         anchor="w").grid(row=0,column=idx,sticky="w",padx=10,pady=(8,0))
            var=tk.StringVar(); self._filter_vars[key]=var
            if opts:
                mb=ctk.CTkOptionMenu(cr,variable=var,values=["(any)"]+opts,fg_color=BG_CARD,
                    button_color=ACCENT2,button_hover_color=ACCENT,text_color=TEXT_MAIN,
                    font=("Courier New",11),width=130)
                mb.set("(any)"); mb.grid(row=1,column=idx,sticky="w",padx=10,pady=(0,8))
            else:
                ctk.CTkEntry(cr,textvariable=var,width=110,font=("Courier New",11),
                    fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=BORDER,
                    placeholder_text="—").grid(row=1,column=idx,sticky="w",padx=10,pady=(0,8))

        sf=ctk.CTkFrame(top,fg_color="transparent"); sf.pack(fill="x",pady=(6,0))
        self._sl(sf,"SEQUENCE")
        sc=ctk.CTkFrame(sf,fg_color=BG_FIELD,corner_radius=8); sc.pack(fill="x")
        self._fseq_mode=tk.StringVar(value="all")
        rr=ctk.CTkFrame(sc,fg_color="transparent"); rr.pack(anchor="w",padx=14,pady=(8,4))
        ctk.CTkLabel(rr,text="Apply to:",font=("Courier New",10),text_color=TEXT_DIM
                     ).pack(side="left",padx=(0,10))
        for val,txt,fg,hov in [("all","All sequences",ACCENT,ACCENT2),
                                ("selected","Selected range",ACCENT2,ACCENT)]:
            ctk.CTkRadioButton(rr,text=txt,variable=self._fseq_mode,value=val,
                font=("Courier New",11),text_color=TEXT_MAIN,fg_color=fg,hover_color=hov,
                command=self._toggle_fseq).pack(side="left",padx=(0,18))
        self._fseq_rf=ctk.CTkFrame(sc,fg_color="transparent")
        self._fseq_rf.pack(anchor="w",padx=14,pady=(0,10))
        self._fseq_s=tk.StringVar(); self._fseq_e=tk.StringVar()
        for lbl,var,ph in [("From",self._fseq_s,"e.g. 0617"),("To  ",self._fseq_e,"e.g. 0703")]:
            ctk.CTkLabel(self._fseq_rf,text=lbl,font=("Courier New",10),
                         text_color=TEXT_DIM).pack(side="left",padx=(0,4))
            ctk.CTkEntry(self._fseq_rf,textvariable=var,width=100,font=("Courier New",11),
                fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=BORDER,
                placeholder_text=ph).pack(side="left",padx=(0,14))
        self._toggle_fseq()

        self._filter_ext=tk.StringVar(value="both")
        er=ctk.CTkFrame(top,fg_color="transparent"); er.pack(fill="x",pady=(8,0))
        ctk.CTkLabel(er,text="Type:",font=("Courier New",10),text_color=TEXT_DIM
                     ).pack(side="left",padx=(0,8))
        for val,txt in [("both","Color + Depth"),("jpg","Color only (.jpg)"),
                         ("png","Depth only (.png)")]:
            ctk.CTkRadioButton(er,text=txt,variable=self._filter_ext,value=val,
                font=("Courier New",11),text_color=TEXT_MAIN,fg_color=ACCENT,
                hover_color=ACCENT2).pack(side="left",padx=12)

        br=ctk.CTkFrame(top,fg_color="transparent"); br.pack(fill="x",pady=(10,0))
        ctk.CTkButton(br,text="⊞  Search",height=38,width=140,fg_color=ACCENT,
            hover_color=ACCENT2,font=("Courier New",12,"bold"),
            command=self._run_filter).pack(side="left")
        self._filter_count=tk.StringVar()
        ctk.CTkLabel(br,textvariable=self._filter_count,font=("Courier New",11),
                     text_color=SUCCESS).pack(side="left",padx=16)
        ctk.CTkButton(br,text="⊡  Export list",height=38,width=130,fg_color=BG_FIELD,
            border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",11),text_color=ACCENT,
            command=self._export_filter_list).pack(side="left",padx=8)
        ctk.CTkButton(br,text="⊠  Copy matched",height=38,width=150,fg_color=BG_FIELD,
            border_width=1,border_color=ACCENT2,hover_color=BG_MID,
            font=("Courier New",11),text_color=ACCENT2,
            command=self._copy_matched).pack(side="left",padx=4)

        tf=ctk.CTkFrame(parent,fg_color=BG_FIELD,corner_radius=0)
        tf.pack(fill="both",expand=True,padx=12,pady=(4,12))
        style=ttk.Style(); style.theme_use("clam")
        style.configure("DS.Treeview",background=BG_FIELD,foreground=TEXT_MAIN,
            fieldbackground=BG_FIELD,rowheight=24,font=("Courier New",10))
        style.configure("DS.Treeview.Heading",background=BG_MID,foreground=ACCENT,
            font=("Courier New",9,"bold"),relief="flat")
        style.map("DS.Treeview",background=[("selected",BG_MID)],foreground=[("selected",ACCENT)])
        style.map("DS.Treeview.Heading",background=[("active",BG_CARD)])
        cols=("base","room","height","angle","distance","lighting","seq","has")
        self._filter_tree=ttk.Treeview(tf,columns=cols,show="headings",
            style="DS.Treeview",selectmode="browse")
        for cid,hdr,w in [("base","Base filename",320),("room","Room",70),
            ("height","Height",60),("angle","Angle",80),("distance","Distance",75),
            ("lighting","Lighting",70),("seq","Seq",50),("has","Has",60)]:
            self._filter_tree.heading(cid,text=hdr.upper())
            self._filter_tree.column(cid,width=w,minwidth=40,anchor="w")
        vsb=ttk.Scrollbar(tf,orient="vertical",command=self._filter_tree.yview)
        self._filter_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right",fill="y"); self._filter_tree.pack(fill="both",expand=True)
        self._filter_tree.tag_configure("even",background=BG_FIELD)
        self._filter_tree.tag_configure("odd",background=BG_CARD)
        self._filter_tree.bind("<Double-1>",self._on_tree_dbl)
        self._filter_tree.bind("<Return>",self._on_tree_dbl)
        self._filter_matches=[]

    def _toggle_fseq(self):
        s="normal" if self._fseq_mode.get()=="selected" else "disabled"
        for c in self._fseq_rf.winfo_children(): c.configure(state=s)

    def _on_tree_dbl(self,event=None):
        sel=self._filter_tree.selection()
        if not sel: return
        idx=0
        for t in self._filter_tree.item(sel[0],"tags"):
            if t.startswith("idx:"): idx=int(t[4:]); break
        self._open_image_picker(idx)

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 3 — MOVE BY ROOM
    # ══════════════════════════════════════════════════════════════════════════
    def _build_move_tab(self,parent):
        parent.configure(fg_color=BG_CARD)
        L=ctk.CTkFrame(parent,fg_color=BG_CARD,corner_radius=0)
        L.pack(side="left",fill="both",expand=True,padx=(12,6),pady=12)
        R=ctk.CTkFrame(parent,fg_color=BG_CARD,corner_radius=0,width=300)
        R.pack(side="right",fill="y",padx=(0,12),pady=12); R.pack_propagate(False)

        self._sl(L,"SELECT  ROOM(S)  TO  MOVE  (FFRRRR)")
        rc=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); rc.pack(fill="x",pady=(6,0))
        r1=ctk.CTkFrame(rc,fg_color="transparent"); r1.pack(fill="x",padx=14,pady=8)
        ctk.CTkLabel(r1,text="Room code (FFRRRR):",font=("Courier New",10),
                     text_color=TEXT_DIM).pack(side="left",padx=(0,10))
        self._move_room_entry=ctk.CTkEntry(r1,width=140,font=("Courier New",12,"bold"),
            fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=ACCENT,
            placeholder_text="e.g. 070701"); self._move_room_entry.pack(side="left",padx=(0,10))
        ctk.CTkButton(r1,text="+ Add",width=80,fg_color=ACCENT2,hover_color=ACCENT,
            font=("Courier New",11,"bold"),command=self._add_move_room).pack(side="left",padx=4)
        ctk.CTkButton(r1,text="Scan rooms",width=110,fg_color=BG_CARD,border_width=1,
            border_color=ACCENT,hover_color=BG_MID,font=("Courier New",11),text_color=ACCENT,
            command=self._scan_rooms).pack(side="left",padx=4)

        self._sl(L,"ROOMS  QUEUED  FOR  MOVE")
        self._rooms_lb=ctk.CTkTextbox(L,height=100,font=("Courier New",11),
            fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=BORDER,corner_radius=6)
        self._rooms_lb.pack(fill="x",pady=(4,0)); self._rooms_lb.configure(state="disabled")
        self._move_rooms=[]
        ctk.CTkButton(L,text="✕  Clear all rooms",width=140,fg_color=BG_FIELD,
            border_width=1,border_color=DANGER,hover_color=BG_MID,
            font=("Courier New",11),text_color=DANGER,
            command=self._clear_move_rooms).pack(anchor="w",pady=(4,0))

        self._sl(L,"DESTINATION  STRUCTURE")
        dc=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); dc.pack(fill="x",pady=(4,0))
        self._move_struct=tk.StringVar(value="room_folder")
        for val,txt,desc in [
            ("room_folder","One folder per room","dest/070701/color/…  &  dest/070701/depth_raw/…"),
            ("flat","Flat — all files together","dest/070701_0.8m_1_close_dim_0617.jpg"),
            ("mirror","Mirror original structure","Keeps relative sub-folders as-is")]:
            rb=ctk.CTkFrame(dc,fg_color="transparent"); rb.pack(anchor="w",padx=14,pady=4)
            ctk.CTkRadioButton(rb,text=txt,variable=self._move_struct,value=val,
                font=("Courier New",11),text_color=TEXT_MAIN,fg_color=ACCENT,
                hover_color=ACCENT2).pack(side="left")
            ctk.CTkLabel(rb,text=f"  ↳ {desc}",font=("Courier New",9),
                         text_color=TEXT_DIM).pack(side="left",padx=6)

        oc=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); oc.pack(fill="x",pady=(10,0))
        self._move_copy=tk.BooleanVar(value=False)
        ctk.CTkCheckBox(oc,text="Copy instead of move  (keep originals in place)",
            variable=self._move_copy,font=("Courier New",11),text_color=TEXT_MAIN,
            fg_color=ACCENT,hover_color=ACCENT2).pack(anchor="w",padx=14,pady=10)

        self._sl(R,"DESTINATION  FOLDER")
        self._move_dest=tk.StringVar()
        ctk.CTkEntry(R,textvariable=self._move_dest,font=("Courier New",10),
            fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=BORDER,
            placeholder_text="Click Browse…").pack(fill="x",pady=(4,6))
        ctk.CTkButton(R,text="Browse destination…",height=36,fg_color=BG_FIELD,
            border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",11),text_color=ACCENT,
            command=self._browse_move_dest).pack(fill="x",pady=4)
        ctk.CTkButton(R,text="⟳  Preview Move",height=38,fg_color=BG_FIELD,
            border_width=1,border_color=ACCENT2,hover_color=BG_MID,
            font=("Courier New",12,"bold"),text_color=ACCENT2,
            command=self._preview_move).pack(fill="x",pady=(20,6))
        ctk.CTkButton(R,text="⇄  Execute Move",height=44,fg_color=ACCENT2,
            hover_color=ACCENT,font=("Courier New",13,"bold"),text_color="white",
            command=self._execute_move).pack(fill="x",pady=6)

        self._sl(L,"MOVE  LOG")
        self._move_log=ctk.CTkTextbox(L,height=160,font=("Courier New",10),
            fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=BORDER,corner_radius=6)
        self._move_log.pack(fill="both",expand=True,pady=(4,0))

    def _add_move_room(self):
        code=self._move_room_entry.get().strip()
        if not re.fullmatch(r"\d{6}",code):
            messagebox.showerror("Invalid","Room code must be exactly 6 digits (FFRRRR)."); return
        if code in self._move_rooms:
            messagebox.showinfo("Duplicate",f"{code} is already in the list."); return
        self._move_rooms.append(code); self._refresh_rooms_list()
        self._move_room_entry.delete(0,"end")

    def _scan_rooms(self):
        root=self._get_root()
        if not root: return
        found=set()
        for _,_,f in walk_images(root):
            p=parse_filename(f)
            if p: found.add(p["room"])
        if not found: messagebox.showinfo("None found","No recognisable images."); return
        win=ctk.CTkToplevel(self); win.title("Select rooms")
        win.geometry("340x420"); win.configure(fg_color=BG_DARK); win.grab_set()
        ctk.CTkLabel(win,text="Rooms found:",font=("Courier New",11,"bold"),
                     text_color=ACCENT).pack(pady=(14,6))
        checks={}; sc=ctk.CTkScrollableFrame(win,fg_color=BG_FIELD,corner_radius=8)
        sc.pack(fill="both",expand=True,padx=16,pady=4)
        for code in sorted(found):
            v=tk.BooleanVar(value=code not in self._move_rooms); checks[code]=v
            ctk.CTkCheckBox(sc,text=code,variable=v,font=("Courier New",11),
                text_color=TEXT_MAIN,fg_color=ACCENT,hover_color=ACCENT2).pack(anchor="w",pady=3)
        def _ok():
            for c,v in checks.items():
                if v.get() and c not in self._move_rooms: self._move_rooms.append(c)
            self._refresh_rooms_list(); win.destroy()
        ctk.CTkButton(win,text="Add selected",fg_color=ACCENT,hover_color=ACCENT2,
                      font=("Courier New",11,"bold"),command=_ok).pack(fill="x",padx=16,pady=10)

    def _clear_move_rooms(self): self._move_rooms.clear(); self._refresh_rooms_list()
    def _refresh_rooms_list(self):
        self._rooms_lb.configure(state="normal"); self._rooms_lb.delete("1.0","end")
        for r in self._move_rooms: self._rooms_lb.insert("end",f"  {r}\n")
        self._rooms_lb.configure(state="disabled")
    def _browse_move_dest(self):
        d=filedialog.askdirectory(); 
        if d: self._move_dest.set(d)
    def _gather_move_plan(self):
        root=self._get_root()
        if not root: return None
        dest=self._move_dest.get().strip()
        if not dest: messagebox.showerror("Error","Please select a destination folder."); return None
        if not self._move_rooms: messagebox.showerror("Error","No rooms selected."); return None
        struct=self._move_struct.get(); plan=[]
        for ap,rel,fname in walk_images(root):
            p=parse_filename(fname)
            if not p or p["room"] not in self._move_rooms: continue
            if struct=="room_folder":
                sub="color" if fname.lower().endswith(".jpg") else "depth_raw"
                dst=os.path.join(dest,p["room"],sub,fname)
            elif struct=="flat": dst=os.path.join(dest,fname)
            else: dst=os.path.join(dest,rel,fname)
            plan.append((ap,dst))
        return plan
    def _preview_move(self):
        plan=self._gather_move_plan()
        if plan is None: return
        self._clr(self._move_log)
        if not plan: self._log(self._move_log,"No matching files found."); return
        jpg=sum(1 for s,_ in plan if s.lower().endswith(".jpg"))
        self._log(self._move_log,f"Found {len(plan)} files: {jpg} color + {len(plan)-jpg} depth")
        self._log(self._move_log,"─"*120)
        for s,d in plan:
            t="[color]" if s.lower().endswith(".jpg") else "[depth]"
            self._log(self._move_log,f"  {t}  {os.path.basename(s):<48}  →  {d}")
        self._log(self._move_log,f"\n{len(plan)} file(s) would be {'copied' if self._move_copy.get() else 'moved'}.")
    def _execute_move(self):
        plan=self._gather_move_plan()
        if plan is None: return
        if not plan: messagebox.showinfo("Nothing to do","No matching files."); return
        verb="copy" if self._move_copy.get() else "move"
        jpg=sum(1 for s,_ in plan if s.lower().endswith(".jpg"))
        if not messagebox.askyesno("Confirm",
            f"{verb.capitalize()} {len(plan)} file(s)?\n  • {jpg} color\n  • {len(plan)-jpg} depth\n\nRooms: {', '.join(self._move_rooms)}"): return
        self._clr(self._move_log); done=errors=0
        for s,d in plan:
            try:
                os.makedirs(os.path.dirname(d),exist_ok=True)
                (shutil.copy2 if self._move_copy.get() else shutil.move)(s,d)
                t="[color]" if s.lower().endswith(".jpg") else "[depth]"
                self._log(self._move_log,f"✔  {t}  {os.path.basename(s)}"); done+=1
            except Exception as e:
                self._log(self._move_log,f"✘  {os.path.basename(s)}  ERROR: {e}"); errors+=1
        summary=f"Done: {done} {verb}d, {errors} error(s)."
        self._log(self._move_log,"\n"+summary); self._set_status(summary)
        messagebox.showinfo("Complete",summary)

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 4 — RESEQUENCE
    # ══════════════════════════════════════════════════════════════════════════
    def _build_resequence_tab(self,parent):
        parent.configure(fg_color=BG_CARD)
        L=ctk.CTkFrame(parent,fg_color=BG_CARD,corner_radius=0)
        L.pack(side="left",fill="both",expand=True,padx=(12,6),pady=12)
        R=ctk.CTkFrame(parent,fg_color=BG_CARD,corner_radius=0,width=300)
        R.pack(side="right",fill="y",padx=(0,12),pady=12); R.pack_propagate(False)

        # ── How it works explainer ────────────────────────────────────────────
        info=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8)
        info.pack(fill="x",pady=(0,8))
        ctk.CTkLabel(info,
            text=("⟳  RESEQUENCE  —  renumbers images so they start at 0001 with no gaps.\n"
                  "Use this after merging files from different capture sessions into one folder.\n"
                  "Sequences are renumbered per group (room + height + angle + distance + lighting)."),
            font=("Courier New",9),text_color=TEXT_DIM,justify="left",anchor="w"
            ).pack(anchor="w",padx=14,pady=10)

        # ── Target folder ─────────────────────────────────────────────────────
        self._sl(L,"TARGET  FOLDER  (folder containing the images to resequence)")
        tf_row=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); tf_row.pack(fill="x",pady=(4,0))
        inner=ctk.CTkFrame(tf_row,fg_color="transparent"); inner.pack(fill="x",padx=14,pady=8)
        self._reseq_folder=tk.StringVar()
        ctk.CTkEntry(inner,textvariable=self._reseq_folder,font=("Courier New",11),
            fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=ACCENT,
            placeholder_text="Browse or type path…",width=420).pack(side="left",padx=(0,10))
        ctk.CTkButton(inner,text="Browse…",width=100,fg_color=ACCENT,hover_color=ACCENT2,
            font=("Courier New",11,"bold"),command=self._browse_reseq_folder
            ).pack(side="left")
        ctk.CTkButton(inner,text="Use dataset root",width=130,fg_color=BG_CARD,
            border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",10),text_color=ACCENT,
            command=lambda: self._reseq_folder.set(self.dataset_path.get().strip())
            ).pack(side="left",padx=8)

        # ── Options ───────────────────────────────────────────────────────────
        self._sl(L,"OPTIONS")
        oc=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); oc.pack(fill="x",pady=(4,0))
        self._reseq_start=tk.IntVar(value=1)   # start sequence at this number
        self._reseq_backup=tk.BooleanVar(value=True)
        self._reseq_dry=tk.BooleanVar(value=True)

        r1=ctk.CTkFrame(oc,fg_color="transparent"); r1.pack(anchor="w",padx=14,pady=(10,4))
        ctk.CTkLabel(r1,text="Start sequence at:",font=("Courier New",10),
                     text_color=TEXT_DIM).pack(side="left",padx=(0,8))
        ctk.CTkEntry(r1,textvariable=self._reseq_start,width=70,font=("Courier New",11),
            fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=ACCENT).pack(side="left")
        ctk.CTkLabel(r1,text="(usually 1 → becomes 0001)",font=("Courier New",9),
                     text_color=TEXT_DIM).pack(side="left",padx=8)

        ctk.CTkCheckBox(oc,text="Dry run  (preview only — no files renamed)",
            variable=self._reseq_dry,font=("Courier New",11),text_color=TEXT_MAIN,
            fg_color=ACCENT,hover_color=ACCENT2).pack(anchor="w",padx=14,pady=4)
        ctk.CTkCheckBox(oc,text="Backup originals  (.bak copy before rename)",
            variable=self._reseq_backup,font=("Courier New",11),text_color=TEXT_MAIN,
            fg_color=ACCENT,hover_color=ACCENT2).pack(anchor="w",padx=14,pady=(4,12))

        # ── Group filter (optional) ────────────────────────────────────────────
        self._sl(L,"RESEQUENCE  ONLY  MATCHING  GROUP  (leave (any) to resequence all)")
        gf=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8); gf.pack(fill="x",pady=(4,0))
        self._reseq_filter={}
        for col,(lbl,key,opts) in enumerate([
            ("Room","rseq_room",None),("Height","rseq_height",HEIGHT_OPTS),
            ("Angle","rseq_angle",ANGLE_OPTS),("Distance","rseq_dist",DIST_OPTS),
            ("Lighting","rseq_light",LIGHT_OPTS)]):
            ctk.CTkLabel(gf,text=lbl,font=("Courier New",10),text_color=TEXT_DIM,
                         anchor="w").grid(row=0,column=col,sticky="w",padx=10,pady=(8,0))
            v=tk.StringVar()
            self._reseq_filter[key]=v
            if opts:
                mb=ctk.CTkOptionMenu(gf,variable=v,values=["(any)"]+opts,fg_color=BG_CARD,
                    button_color=ACCENT2,button_hover_color=ACCENT,text_color=TEXT_MAIN,
                    font=("Courier New",11),width=120)
                mb.set("(any)"); mb.grid(row=1,column=col,sticky="w",padx=10,pady=(0,8))
            else:
                ctk.CTkEntry(gf,textvariable=v,width=100,font=("Courier New",11),
                    fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=BORDER,
                    placeholder_text="—").grid(row=1,column=col,sticky="w",padx=10,pady=(0,8))

        # ── Buttons & log ─────────────────────────────────────────────────────
        ctk.CTkButton(R,text="⟳  Preview Resequence",height=38,fg_color=BG_FIELD,
            border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",12,"bold"),text_color=ACCENT,
            command=lambda:self._run_resequence(dry=True)).pack(fill="x",pady=(30,6))
        ctk.CTkButton(R,text="✔  Apply Resequence",height=44,fg_color=ACCENT,
            hover_color=ACCENT2,font=("Courier New",13,"bold"),text_color="white",
            command=lambda:self._run_resequence(dry=False)).pack(fill="x",pady=6)

        self._reseq_summary=tk.StringVar(value="")
        ctk.CTkLabel(R,textvariable=self._reseq_summary,font=("Courier New",10),
                     text_color=SUCCESS,justify="left").pack(anchor="w",padx=8,pady=(12,0))

        # ── .bak restorer ─────────────────────────────────────────────────────
        sep=ctk.CTkFrame(L,fg_color=BORDER,height=1,corner_radius=0)
        sep.pack(fill="x",pady=(14,0))
        self._sl(L,"BULK  .BAK  →  ORIGINAL  EXTENSION  RESTORER")
        bak_info=ctk.CTkFrame(L,fg_color=BG_FIELD,corner_radius=8)
        bak_info.pack(fill="x",pady=(4,0))
        ctk.CTkLabel(bak_info,
            text="Finds all .bak files and renames them back to their original extension.  "
                 "e.g.  file.jpg.bak  ->  file.jpg",
            font=("Courier New",9),text_color=TEXT_DIM,justify="left",anchor="w"
            ).pack(anchor="w",padx=14,pady=8)

        bak_row=ctk.CTkFrame(L,fg_color="transparent"); bak_row.pack(fill="x",pady=(6,0))
        self._bak_folder=tk.StringVar()
        ctk.CTkEntry(bak_row,textvariable=self._bak_folder,font=("Courier New",11),
            fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=ACCENT,
            placeholder_text="Select folder containing .bak files…",width=440
            ).pack(side="left",padx=(0,10))
        ctk.CTkButton(bak_row,text="Browse…",width=90,fg_color=ACCENT,hover_color=ACCENT2,
            font=("Courier New",11,"bold"),
            command=lambda:self._bak_folder.set(
                filedialog.askdirectory(title="Select folder with .bak files") or
                self._bak_folder.get())
            ).pack(side="left",padx=(0,8))
        ctk.CTkButton(bak_row,text="Use dataset root",width=130,fg_color=BG_CARD,
            border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",10),text_color=ACCENT,
            command=lambda:self._bak_folder.set(self.dataset_path.get().strip())
            ).pack(side="left")

        bak_opts=ctk.CTkFrame(L,fg_color="transparent"); bak_opts.pack(fill="x",pady=(6,0))
        self._bak_dry=tk.BooleanVar(value=True)
        self._bak_delete=tk.BooleanVar(value=False)
        ctk.CTkCheckBox(bak_opts,text="Dry run  (preview only)",
            variable=self._bak_dry,font=("Courier New",11),text_color=TEXT_MAIN,
            fg_color=ACCENT,hover_color=ACCENT2).pack(side="left",padx=(0,16))
        ctk.CTkCheckBox(bak_opts,text="Delete .bak after restoring",
            variable=self._bak_delete,font=("Courier New",11),text_color=TEXT_MAIN,
            fg_color=DANGER,hover_color="#C94040").pack(side="left")

        bak_btns=ctk.CTkFrame(L,fg_color="transparent"); bak_btns.pack(fill="x",pady=(6,0))
        ctk.CTkButton(bak_btns,text="⟳  Preview Restore",height=36,width=160,
            fg_color=BG_FIELD,border_width=1,border_color=ACCENT,hover_color=BG_MID,
            font=("Courier New",11,"bold"),text_color=ACCENT,
            command=lambda:self._run_bak_restore(dry=True)).pack(side="left",padx=(0,8))
        ctk.CTkButton(bak_btns,text="✔  Apply Restore",height=36,width=150,
            fg_color=ACCENT,hover_color=ACCENT2,
            font=("Courier New",11,"bold"),text_color="white",
            command=lambda:self._run_bak_restore(dry=False)).pack(side="left")

        sep2=ctk.CTkFrame(L,fg_color=BORDER,height=1,corner_radius=0)
        sep2.pack(fill="x",pady=(12,0))
        self._sl(L,"LOG")
        self._reseq_log=ctk.CTkTextbox(L,font=("Courier New",10),fg_color=BG_FIELD,
            text_color=TEXT_MAIN,border_color=BORDER,corner_radius=6)
        self._reseq_log.pack(fill="both",expand=True,pady=(4,0))

    def _browse_reseq_folder(self):
        d=filedialog.askdirectory(title="Select folder to resequence")
        if d: self._reseq_folder.set(d)

    def _run_resequence(self,dry=True):
        folder=self._reseq_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Error","Please select a valid target folder."); return
        try: start=int(self._reseq_start.get())
        except: messagebox.showerror("Error","Start sequence must be an integer."); return
        if start<1: messagebox.showerror("Error","Start sequence must be ≥ 1."); return

        rf=self._reseq_filter
        def _matches_filter(p):
            for field,key in [("room","rseq_room"),("height","rseq_height"),
                               ("angle","rseq_angle"),("distance","rseq_dist"),
                               ("lighting","rseq_light")]:
                v=rf[key].get().strip()
                if v and v!="(any)" and p[field]!=v: return False
            return True

        self._clr(self._reseq_log)
        self._reseq_summary.set("Scanning…")
        verb="DRY RUN" if dry else "APPLYING"
        self._log(self._reseq_log,f"{'─'*6}  {verb}  RESEQUENCE  {'─'*6}")
        self._log(self._reseq_log,f"Folder: {folder}")
        self._log(self._reseq_log,f"Start at: {start:04d}  |  Dry run: {dry}\n")

        def _worker():
            # ── 1. Collect all images, group by (room,height,angle,dist,light) ──
            groups: dict[str, list[dict]] = {}   # group_key -> list of file-info dicts
            for abs_path, rel, fname in walk_images(folder):
                p = parse_filename(fname)
                if not p: continue
                if not _matches_filter(p): continue
                gk = group_key(p)
                if gk not in groups: groups[gk] = []
                groups[gk].append({
                    "abs": abs_path, "parts": p, "fname": fname,
                    "dir": os.path.dirname(abs_path)
                })

            if not groups:
                self.after(0,lambda:self._reseq_summary.set("No matching images found."))
                self.after(0,lambda:self._log(self._reseq_log,"No matching images found."))
                return

            # ── 2. Collect ALL sequence slots across every group, sort globally ──
            # Each "slot" = one (old_seq_number, group_key) pair.
            # We sort ALL slots together by their current sequence number so the
            # final numbering is continuous across groups (no reset per group).
            all_slots = []   # list of (seq_int, gk, seq_str)
            for gk, files in groups.items():
                seen_seqs = set()
                for fi in files:
                    seq = fi["parts"]["sequence"]
                    if seq not in seen_seqs:
                        seen_seqs.add(seq)
                        all_slots.append((int(seq), gk, seq))

            all_slots.sort(key=lambda x: x[0])   # sort by existing sequence number globally

            # Assign new sequence numbers in one continuous run
            slot_new_seq: dict[tuple, str] = {}   # (gk, old_seq_str) -> new_seq_str
            next_seq = start
            for seq_int, gk, seq_str in all_slots:
                slot_new_seq[(gk, seq_str)] = f"{next_seq:04d}"
                next_seq += 1

            lines=[]; total_renamed=0; total_skipped=0; errors=0
            all_rename_ops = []   # collect all ops; apply in one 2-pass batch

            for gk in sorted(groups.keys()):
                files = groups[gk]
                # Pair color+depth by sequence
                seq_pairs: dict[str, dict] = {}
                for fi in files:
                    seq = fi["parts"]["sequence"]
                    if seq not in seq_pairs: seq_pairs[seq] = {"color": None, "depth": None}
                    if fi["parts"]["is_depth"]: seq_pairs[seq]["depth"] = fi
                    else:                       seq_pairs[seq]["color"] = fi

                sorted_seqs = sorted(seq_pairs.keys(), key=lambda s: int(s))
                new_seqs_for_group = [slot_new_seq[(gk, s)] for s in sorted_seqs]
                lines.append(f"\nGroup: {gk}  ({len(sorted_seqs)} slot(s)  →  "
                              f"{new_seqs_for_group[0]}…{new_seqs_for_group[-1]})")

                # Check if already correctly sequenced (compare actual vs assigned)
                already_ok = all(s == slot_new_seq[(gk, s)] for s in sorted_seqs)
                if already_ok:
                    lines.append(f"  ✔ Already correctly sequenced. Skipping.")
                    total_skipped += len(sorted_seqs)
                    continue

                for old_seq in sorted_seqs:
                    new_seq_str = slot_new_seq[(gk, old_seq)]
                    pair = seq_pairs[old_seq]
                    for role in ("color", "depth"):
                        fi = pair[role]
                        if fi is None: continue
                        new_name = build_filename({**fi["parts"], "sequence": new_seq_str})
                        new_abs  = os.path.join(fi["dir"], new_name)
                        temp_abs = fi["abs"] + ".reseq_tmp"
                        if fi["abs"] == new_abs:
                            lines.append(f"  = {fi['fname']}  (unchanged)")
                            total_skipped += 1
                            continue
                        lines.append(f"  {fi['fname']}  →  {new_name}")
                        all_rename_ops.append((fi["abs"], new_abs, temp_abs))
                        total_renamed += 1

            # ── Single 2-pass batch rename (avoids collisions across groups) ──
            if not dry and all_rename_ops:
                # Pass 1: every file → .reseq_tmp
                for old, new, tmp in all_rename_ops:
                    try:
                        if self._reseq_backup.get(): shutil.copy2(old, old+".bak")
                        os.rename(old, tmp)
                    except Exception as e:
                        lines.append(f"  ✘ TEMP ERROR: {os.path.basename(old)}: {e}")
                        errors += 1
                # Pass 2: .reseq_tmp → final name
                for old, new, tmp in all_rename_ops:
                    try: os.rename(tmp, new)
                    except Exception as e:
                        lines.append(f"  ✘ FINAL ERROR: {os.path.basename(tmp)}: {e}")
                        errors += 1

            bulk = "\n".join(lines)
            summary = (f"{'Preview: ' if dry else ''}"
                       f"{total_renamed} file(s) {'would be ' if dry else ''}renamed  |  "
                       f"{total_skipped} skipped  |  {errors} error(s)")

            def _update():
                self._log(self._reseq_log, bulk)
                self._log(self._reseq_log, f"\n{'─'*40}\n{summary}")
                self._reseq_summary.set(summary)
                self._set_status(summary)
                if not dry and total_renamed > 0:
                    messagebox.showinfo("Resequence complete", summary)
            self.after(0, _update)

        threading.Thread(target=_worker, daemon=True).start()

    def _run_bak_restore(self, dry=True):
        folder = self._bak_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Error", "Please select a valid folder."); return

        self._clr(self._reseq_log)
        verb = "DRY RUN" if dry else "APPLYING"
        self._log(self._reseq_log, f"{'─'*6}  {verb}  .BAK RESTORE  {'─'*6}")
        self._log(self._reseq_log, "Folder: " + folder)

        def _worker():
            lines = []; restored = skipped = errors = 0
            for dirpath, _, files in os.walk(folder):
                for fname in sorted(files):
                    if not fname.lower().endswith(".bak"): continue
                    bak_path = os.path.join(dirpath, fname)
                    # Strip the .bak suffix to get the original path
                    orig_path = bak_path[:-4]   # removes ".bak"
                    orig_name = os.path.basename(orig_path)

                    if os.path.exists(orig_path):
                        lines.append(f"  SKIP  {fname}  (original already exists)")
                        skipped += 1
                        continue

                    lines.append(f"  {fname}  →  {orig_name}")

                    if not dry:
                        try:
                            os.rename(bak_path, orig_path)
                            if self._bak_delete.get():
                                # already renamed — nothing to delete
                                pass
                            restored += 1
                        except Exception as e:
                            lines.append(f"    ✘ ERROR: {e}")
                            errors += 1
                    else:
                        restored += 1   # count for preview

            bulk = "\n".join(lines) if lines else "  No .bak files found."
            summary = (f"{'Preview: ' if dry else ''}"
                       f"{restored} file(s) {'would be ' if dry else ''}restored  |  "
                       f"{skipped} skipped (original exists)  |  {errors} error(s)")

            def _update():
                self._log(self._reseq_log, bulk)
                self._log(self._reseq_log, "\n" + "─"*40 + "\n" + summary)
                self._reseq_summary.set(summary)
                self._set_status(summary)
                if not dry and restored > 0:
                    messagebox.showinfo(".bak Restore complete", summary)
            self.after(0, _update)

        threading.Thread(target=_worker, daemon=True).start()

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 5 — PREVIEW
    # ══════════════════════════════════════════════════════════════════════════
    def _build_preview_tab(self,parent):
        parent.configure(fg_color=BG_CARD)
        ctrl=ctk.CTkFrame(parent,fg_color=BG_CARD); ctrl.pack(fill="x",padx=12,pady=12)
        ctk.CTkButton(ctrl,text="◈  Scan & Preview All Files",height=38,width=220,
            fg_color=ACCENT,hover_color=ACCENT2,font=("Courier New",12,"bold"),
            command=self._scan_all).pack(side="left")
        self._scan_count=tk.StringVar()
        ctk.CTkLabel(ctrl,textvariable=self._scan_count,font=("Courier New",11),
                     text_color=SUCCESS).pack(side="left",padx=16)
        hdr=ctk.CTkFrame(parent,fg_color=BG_MID,corner_radius=0,height=24)
        hdr.pack(fill="x",padx=12); hdr.pack_propagate(False)
        for txt,w in [("Path",35),("Room",8),("Height",7),("Angle",10),
                       ("Distance",9),("Lighting",9),("Seq",6),("Type",6)]:
            ctk.CTkLabel(hdr,text=txt,font=("Courier New",9,"bold"),text_color=ACCENT,
                         width=w*7,anchor="w").pack(side="left",padx=4)
        self._preview_box=ctk.CTkTextbox(parent,font=("Courier New",10),fg_color=BG_FIELD,
            text_color=TEXT_MAIN,border_color=BORDER,corner_radius=0)
        self._preview_box.pack(fill="both",expand=True,padx=12,pady=(0,12))

    # ══════════════════════════════════════════════════════════════════════════
    #  RENAME LOGIC
    # ══════════════════════════════════════════════════════════════════════════
    def _gather_rename_plan(self):
        root=self._get_root()
        if not root: return []
        rv=self._rename_vars; rfv=self._rfilter_vars
        use_range=self._seq_mode.get()=="selected"
        ss=self._seq_s.get().strip() if use_range else ""
        se=self._seq_e.get().strip() if use_range else ""
        plan=[]; seen=set()
        for ap,_,fname in walk_images(root):
            p=parse_filename(fname)
            if not p: continue
            skip=False
            for field,key in [("room","rf_room"),("height","rf_height"),
                               ("angle","rf_angle"),("distance","rf_dist"),("lighting","rf_light")]:
                v=rfv[key].get()
                if v and v!="(any)" and p[field]!=v: skip=True; break
            if skip: continue
            if ss and int(p["sequence"])<int(ss): continue
            if se and int(p["sequence"])>int(se): continue
            new=dict(p)
            for field,nk in [("room","new_room"),("height","new_height"),
                              ("angle","new_angle"),("distance","new_dist"),("lighting","new_light")]:
                val=rv[nk].get()
                if val and val!="(keep)": new[field]=val
            nn=build_filename(new); na=os.path.join(os.path.dirname(ap),nn)
            if nn==fname or na in seen: continue
            seen.add(na); plan.append((ap,na))
        return plan

    def _preview_rename(self):
        plan=self._gather_rename_plan(); self._clr(self._rename_log)
        if not plan: self._log(self._rename_log,"No matching files found."); return
        self._log(self._rename_log,f"{'OLD':<55}  →  NEW"); self._log(self._rename_log,"─"*110)
        for old,new in plan:
            self._log(self._rename_log,f"  {os.path.basename(old):<53}  →  {os.path.basename(new)}")
        self._log(self._rename_log,f"\n{len(plan)} file(s) would be renamed.")

    def _apply_rename(self):
        plan=self._gather_rename_plan()
        if not plan: messagebox.showinfo("Nothing to do","No matching files."); return
        if not messagebox.askyesno("Confirm",f"Rename {len(plan)} file(s)?"): return
        self._clr(self._rename_log); renamed=errors=0
        for old,new in plan:
            try:
                if self._backup.get(): shutil.copy2(old,old+".bak")
                os.rename(old,new)
                self._log(self._rename_log,f"✔  {os.path.basename(old)}  →  {os.path.basename(new)}")
                renamed+=1
            except Exception as e:
                self._log(self._rename_log,f"✘  {os.path.basename(old)}  ERROR: {e}"); errors+=1
        summary=f"Done: {renamed} renamed, {errors} error(s)."
        self._log(self._rename_log,"\n"+summary); self._set_status(summary)
        messagebox.showinfo("Complete",summary)

    # ══════════════════════════════════════════════════════════════════════════
    #  FILTER LOGIC
    # ══════════════════════════════════════════════════════════════════════════
    def _run_filter(self):
        if not self._get_root(): return
        self._filter_count.set("Searching…")
        threading.Thread(target=self._run_filter_worker,daemon=True).start()

    def _run_filter_worker(self):
        root=self.dataset_path.get().strip()
        if not root or not os.path.isdir(root): return
        fv=self._filter_vars; ext=self._filter_ext.get()
        ur=self._fseq_mode.get()=="selected"
        ss=self._fseq_s.get().strip() if ur else ""
        se=self._fseq_e.get().strip() if ur else ""
        grouped={}
        for ap,_,fname in walk_images(root):
            p=parse_filename(fname)
            if not p: continue
            skip=False
            for field,key in [("room","f_room"),("height","f_height"),
                               ("angle","f_angle"),("distance","f_dist"),("lighting","f_light")]:
                v=fv[key].get().strip()
                if v and v!="(any)" and p[field]!=v: skip=True; break
            if skip: continue
            if ss and int(p["sequence"])<int(ss): continue
            if se and int(p["sequence"])>int(se): continue
            bk=base_key(p)
            if bk not in grouped:
                grouped[bk]={"base_key":bk,"parts":p,"color_path":None,"depth_path":None}
            if p["ext"]==".jpg": grouped[bk]["color_path"]=ap
            else:                grouped[bk]["depth_path"]=ap
        matches=[r for r in grouped.values()
                 if not(ext=="jpg" and not r["color_path"])
                 and not(ext=="png" and not r["depth_path"])]
        matches.sort(key=lambda r:int(r["parts"]["sequence"]))
        self.after(0,lambda m=matches:self._populate_tree(m))

    def _populate_tree(self,matches):
        self._filter_matches=matches; tree=self._filter_tree
        tree.delete(*tree.get_children())
        for idx,rec in enumerate(matches):
            p=rec["parts"]
            has="+".join(("JPG",)*bool(rec["color_path"])+("PNG",)*bool(rec["depth_path"]))
            tree.insert("","end",
                values=(rec["base_key"],p["room"],p["height"],
                        ANGLE_LABEL.get(p["angle"],p["angle"]),p["distance"],
                        LIGHT_LABEL.get(p["lighting"],p["lighting"]),p["sequence"],has),
                tags=("even" if idx%2==0 else "odd",f"idx:{idx}"))
        self._filter_count.set(f"{len(matches)} unique record(s)")
        self._set_status(f"Filter: {len(matches)} unique record(s)")

    def _refresh_tree_row(self,idx):
        ch=self._filter_tree.get_children()
        if idx>=len(ch): return
        rec=self._filter_matches[idx]; p=rec["parts"]
        has="+".join(("JPG",)*bool(rec["color_path"])+("PNG",)*bool(rec["depth_path"]))
        self._filter_tree.item(ch[idx],
            values=(rec["base_key"],p["room"],p["height"],
                    ANGLE_LABEL.get(p["angle"],p["angle"]),p["distance"],
                    LIGHT_LABEL.get(p["lighting"],p["lighting"]),p["sequence"],has),
            tags=("even" if idx%2==0 else "odd",f"idx:{idx}"))

    # ══════════════════════════════════════════════════════════════════════════
    #  RESEQUENCE HELPER (used by image viewer delete)
    # ══════════════════════════════════════════════════════════════════════════
    @staticmethod
    def _resequence_on_disk(deleted_rec,all_matches,deleted_idx,object_counts):
        logs=[]; dsq=int(deleted_rec["parts"]["sequence"])
        gk=group_key(deleted_rec["parts"])
        to_shift=sorted(
            [(i,r) for i,r in enumerate(all_matches)
             if i!=deleted_idx and group_key(r["parts"])==gk
             and int(r["parts"]["sequence"])>dsq],
            key=lambda x:int(x[1]["parts"]["sequence"]))
        for _,rec in to_shift:
            osq=rec["parts"]["sequence"]; nsq=f"{int(osq)-1:04d}"
            ok=rec["base_key"]; nk=base_key({**rec["parts"],"sequence":nsq}); rok=True
            for pk,isd,ext in [("color_path",False,".jpg"),("depth_path",True,".png")]:
                op=rec[pk]
                if not op or not os.path.isfile(op): continue
                nn=build_filename({**rec["parts"],"sequence":nsq,"is_depth":isd,"ext":ext})
                np=os.path.join(os.path.dirname(op),nn)
                try: os.rename(op,np); rec[pk]=np; logs.append(f"  ↳ {os.path.basename(op)}  →  {nn}")
                except Exception as e: logs.append(f"  ✘ {os.path.basename(op)}: {e}"); rok=False
            if rok:
                rec["parts"]["sequence"]=nsq; rec["base_key"]=nk
                if ok in object_counts: object_counts[nk]=object_counts.pop(ok)
        return logs

    # ══════════════════════════════════════════════════════════════════════════
    #  IMAGE VIEWER
    # ══════════════════════════════════════════════════════════════════════════
    def _open_image_picker(self,start_idx):
        try: from PIL import Image,ImageTk; _pil=True
        except: _pil=False
        matches=self._filter_matches
        if not matches: return
        IMG_W,IMG_H=900,560
        state={"idx":start_idx,"mode":"color","imgtk":None,"resize_job":None,
               "object_counts":{},"obj_cache":None,"save_dir":None}
        win=ctk.CTkToplevel(self); win.title("Image Viewer")
        win.geometry(f"{IMG_W+300}x{IMG_H+160}"); win.minsize(900,600)
        win.resizable(True,True); win.configure(fg_color=BG_DARK)
        win.grab_set(); win.lift(); win.focus_force()

        top=ctk.CTkFrame(win,fg_color=BG_MID,corner_radius=0,height=38)
        top.pack(fill="x"); top.pack_propagate(False)
        counter_lbl=ctk.CTkLabel(top,text="",font=("Courier New",10),text_color=TEXT_DIM); counter_lbl.pack(side="left",padx=14)
        key_lbl=ctk.CTkLabel(top,text="",font=("Courier New",10,"bold"),text_color=TEXT_MAIN); key_lbl.pack(side="left",padx=8)
        detail_lbl=ctk.CTkLabel(top,text="",font=("Courier New",9),text_color=TEXT_DIM); detail_lbl.pack(side="right",padx=16)
        seq_lbl=ctk.CTkLabel(top,text="",font=("Courier New",10),text_color=ACCENT); seq_lbl.pack(side="right",padx=14)

        body=ctk.CTkFrame(win,fg_color=BG_DARK,corner_radius=0); body.pack(fill="both",expand=True)
        cf=ctk.CTkFrame(body,fg_color=BG_FIELD,corner_radius=0); cf.pack(side="left",fill="both",expand=True)
        canvas=tk.Canvas(cf,bg="#0F1117",highlightthickness=0); canvas.pack(fill="both",expand=True)
        nil=ctk.CTkLabel(cf,text="No image available",font=("Courier New",13),text_color=TEXT_DIM,fg_color="transparent")

        rp=ctk.CTkFrame(body,fg_color=BG_CARD,corner_radius=0,width=280)
        rp.pack(side="right",fill="y"); rp.pack_propagate(False)
        rpt=ctk.CTkFrame(rp,fg_color=BG_CARD,corner_radius=0); rpt.pack(fill="x",padx=12,pady=(12,4))
        ctk.CTkLabel(rpt,text="OBJECT  COUNTER",font=("Courier New",9,"bold"),text_color=ACCENT).pack(anchor="w")
        sv=tk.StringVar()
        sf=ctk.CTkFrame(rpt,fg_color=BG_FIELD,corner_radius=6); sf.pack(fill="x",pady=(6,0))
        ctk.CTkLabel(sf,text="⌕",font=("Courier New",13),text_color=TEXT_DIM).pack(side="left",padx=(8,2),pady=4)
        ctk.CTkEntry(sf,textvariable=sv,font=("Courier New",10),fg_color="transparent",
            text_color=TEXT_MAIN,border_width=0,placeholder_text="Search objects…",
            placeholder_text_color=TEXT_DIM).pack(side="left",fill="x",expand=True,pady=4,padx=(0,6))
        csb=ctk.CTkButton(sf,text="×",width=22,height=22,fg_color="transparent",hover_color=BG_MID,
            font=("Courier New",12,"bold"),text_color=TEXT_DIM,command=lambda:sv.set(""))
        ctk.CTkButton(rp,text="+ Add Object",height=32,fg_color=ACCENT2,hover_color=ACCENT,
            font=("Courier New",11,"bold"),text_color="white",
            command=lambda:_add_obj_dlg()).pack(fill="x",padx=12,pady=(4,6))
        os_frame=ctk.CTkScrollableFrame(rp,fg_color=BG_FIELD,corner_radius=6)
        os_frame.pack(fill="both",expand=True,padx=12,pady=(0,6))
        _obj_rows={}; _cnt_vars={}

        def _apply_search(*_):
            q=sv.get().strip().lower()
            if q: csb.pack(side="right",padx=(0,4))
            else: csb.pack_forget()
            for name,(row,_) in _obj_rows.items():
                if not q or q in name.lower(): row.pack(fill="x",pady=2)
                else: row.pack_forget()
        sv.trace_add("write",_apply_search)

        def _all_objs():
            if state["obj_cache"] is None:
                s=set()
                for d in state["object_counts"].values(): s.update(d.keys())
                state["obj_cache"]=sorted(s)
            return state["obj_cache"]
        def _inv_cache(): state["obj_cache"]=None
        def _get_cnts():
            k=matches[state["idx"]]["base_key"]
            if k not in state["object_counts"]: state["object_counts"][k]={}
            return state["object_counts"][k]
        def _rebuild_obj():
            for w in os_frame.winfo_children(): w.destroy()
            _obj_rows.clear(); _cnt_vars.clear()
            cnts=_get_cnts(); objs=_all_objs()
            for n in objs: cnts.setdefault(n,0)
            if not objs:
                ctk.CTkLabel(os_frame,text="No objects added yet.\nClick '+ Add Object'.",
                    font=("Courier New",9),text_color=TEXT_DIM,justify="center").pack(pady=20); return
            for n in objs: _add_obj_row(n,cnts)
            _apply_search()
        def _add_obj_row(name,cnts):
            row=ctk.CTkFrame(os_frame,fg_color=BG_MID,corner_radius=4); row.pack(fill="x",pady=2)
            ctk.CTkLabel(row,text=name,font=("Courier New",10),text_color=TEXT_MAIN,
                         anchor="w",width=86).pack(side="left",padx=(8,2),pady=6)
            cv=tk.StringVar(value=str(cnts.get(name,0))); _cnt_vars[name]=cv; _obj_rows[name]=(row,cv)
            ctk.CTkButton(row,text="−",width=26,height=26,fg_color=BG_FIELD,hover_color=DANGER,
                font=("Courier New",13,"bold"),text_color=TEXT_MAIN,corner_radius=4,
                command=lambda n=name:_chg(n,-1)).pack(side="left",padx=1)
            ctk.CTkLabel(row,textvariable=cv,font=("Courier New",12,"bold"),text_color=ACCENT,
                         width=30,anchor="center").pack(side="left",padx=1)
            ctk.CTkButton(row,text="+",width=26,height=26,fg_color=BG_FIELD,hover_color=SUCCESS,
                font=("Courier New",13,"bold"),text_color=TEXT_MAIN,corner_radius=4,
                command=lambda n=name:_chg(n,+1)).pack(side="left",padx=1)
            ctk.CTkButton(row,text="✎",width=26,height=26,fg_color=BG_FIELD,hover_color=WARNING,
                font=("Courier New",12,"bold"),text_color=WARNING,corner_radius=4,
                command=lambda n=name:_edit_obj(n)).pack(side="left",padx=(3,1))
            ctk.CTkButton(row,text="✕",width=26,height=26,fg_color=BG_FIELD,hover_color=DANGER,
                font=("Courier New",11,"bold"),text_color=DANGER,corner_radius=4,
                command=lambda n=name:_del_obj(n)).pack(side="left",padx=(1,4))
        def _chg(name,delta):
            c=_get_cnts(); c[name]=max(0,c.get(name,0)+delta)
            if name in _cnt_vars: _cnt_vars[name].set(str(c[name]))
        def _del_obj(name):
            if not messagebox.askyesno("Delete object",f"Remove '{name}' from every image?",parent=win): return
            for d in state["object_counts"].values(): d.pop(name,None)
            _inv_cache(); _rebuild_obj()
        def _edit_obj(old):
            dlg=ctk.CTkToplevel(win); dlg.title("Edit Object"); dlg.geometry("320x170")
            dlg.resizable(False,False); dlg.configure(fg_color=BG_DARK); dlg.grab_set(); dlg.lift(); dlg.focus_force()
            ctk.CTkLabel(dlg,text="Rename object:",font=("Courier New",10),text_color=TEXT_DIM).pack(pady=(18,4))
            nv=tk.StringVar(value=old)
            ent=ctk.CTkEntry(dlg,textvariable=nv,width=240,font=("Courier New",12),fg_color=BG_FIELD,text_color=TEXT_MAIN,border_color=ACCENT)
            ent.pack(pady=4); ent.select_range(0,"end"); ent.focus_set()
            ev=tk.StringVar()
            ctk.CTkLabel(dlg,textvariable=ev,font=("Courier New",9),text_color=DANGER).pack(pady=(2,0))
            def _ok():
                new=nv.get().strip()
                if not new: ev.set("Name cannot be empty."); return
                if new==old: dlg.destroy(); return
                if new in _all_objs(): ev.set(f"'{new}' already exists."); return
                for d in state["object_counts"].values():
                    if old in d: d[new]=d.pop(old)
                _inv_cache(); dlg.destroy(); _rebuild_obj()
            ctk.CTkButton(dlg,text="Save",height=36,fg_color=ACCENT,hover_color=ACCENT2,
                font=("Courier New",11,"bold"),text_color="white",command=_ok).pack(fill="x",padx=20,pady=(8,4))
            dlg.bind("<Return>",lambda e:_ok())
        def _add_obj_dlg():
            dlg=ctk.CTkToplevel(win); dlg.title("Add Object"); dlg.geometry("320x160")
            dlg.resizable(False,False); dlg.configure(fg_color=BG_DARK); dlg.grab_set(); dlg.lift(); dlg.focus_force()
            ctk.CTkLabel(dlg,text="Object name:",font=("Courier New",11),text_color=TEXT_MAIN).pack(pady=(20,6))
            ent=ctk.CTkEntry(dlg,width=220,font=("Courier New",12),fg_color=BG_FIELD,
                text_color=TEXT_MAIN,border_color=ACCENT,placeholder_text="e.g. chair")
            ent.pack(pady=4); ent.focus_set()
            def _save():
                name=ent.get().strip()
                if not name: return
                for d in state["object_counts"].values(): d.setdefault(name,0)
                _get_cnts().setdefault(name,0); _inv_cache(); dlg.destroy(); _rebuild_obj()
            ctk.CTkButton(dlg,text="Save",fg_color=ACCENT,hover_color=ACCENT2,
                font=("Courier New",11,"bold"),command=_save).pack(pady=12)
            dlg.bind("<Return>",lambda e:_save())

        bot=ctk.CTkFrame(win,fg_color=BG_MID,corner_radius=0,height=100)
        bot.pack(fill="x",side="bottom"); bot.pack_propagate(False)
        nav=ctk.CTkFrame(bot,fg_color="transparent"); nav.pack(pady=(10,2))
        prev_btn=ctk.CTkButton(nav,text="◀",width=52,height=36,fg_color=BG_FIELD,border_width=1,border_color=BORDER,hover_color=BG_CARD,font=("Courier New",14,"bold"),text_color=TEXT_MAIN); prev_btn.pack(side="left",padx=6)
        color_tab=ctk.CTkButton(nav,text="🖼  Color",width=120,height=36,fg_color=ACCENT,hover_color=ACCENT2,font=("Courier New",11,"bold"),text_color="white"); color_tab.pack(side="left",padx=4)
        depth_tab=ctk.CTkButton(nav,text="◧  Depth",width=120,height=36,fg_color=BG_FIELD,border_width=1,border_color=ACCENT2,hover_color=BG_CARD,font=("Courier New",11),text_color=ACCENT2); depth_tab.pack(side="left",padx=4)
        next_btn=ctk.CTkButton(nav,text="▶",width=52,height=36,fg_color=BG_FIELD,border_width=1,border_color=BORDER,hover_color=BG_CARD,font=("Courier New",14,"bold"),text_color=TEXT_MAIN); next_btn.pack(side="left",padx=6)
        ctk.CTkButton(nav,text="💾  Save Record",width=140,height=36,fg_color=SUCCESS,hover_color="#2AB87A",font=("Courier New",11,"bold"),text_color="white",command=lambda:_save_excel()).pack(side="left",padx=8)
        ctk.CTkButton(nav,text="📂  Open Record",width=140,height=36,fg_color=BG_FIELD,border_width=1,border_color=SUCCESS,hover_color=BG_MID,font=("Courier New",11,"bold"),text_color=SUCCESS,command=lambda:_open_excel_dlg()).pack(side="left",padx=4)
        ctk.CTkButton(nav,text="✎  Rename",width=100,height=36,fg_color=WARNING,hover_color="#D4901E",font=("Courier New",11,"bold"),text_color="white",command=lambda:_rename_dlg()).pack(side="left",padx=4)
        ctk.CTkButton(nav,text="🗑  Delete",width=100,height=36,fg_color=DANGER,hover_color="#C94040",font=("Courier New",11,"bold"),text_color="white",command=lambda:_delete_dlg()).pack(side="left",padx=4)
        fname_lbl=ctk.CTkLabel(bot,text="",font=("Courier New",9),text_color=TEXT_DIM); fname_lbl.pack(pady=(2,6))

        def _render(path):
            nil.place_forget(); canvas.delete("all")
            if not path or not os.path.isfile(path): nil.place(relx=.5,rely=.5,anchor="center"); return
            if not _pil:
                nil.configure(text="Install Pillow:\npip install Pillow"); nil.place(relx=.5,rely=.5,anchor="center"); return
            try:
                img=Image.open(path); cw=canvas.winfo_width() or IMG_W; ch=canvas.winfo_height() or IMG_H
                img.thumbnail((cw,ch),Image.LANCZOS); imgtk=ImageTk.PhotoImage(img)
                state["imgtk"]=imgtk; canvas.create_image(cw//2,ch//2,anchor="center",image=imgtk)
            except Exception as ex:
                nil.configure(text=f"Cannot load image:\n{ex}"); nil.place(relx=.5,rely=.5,anchor="center")

        def _hl_tabs():
            if state["mode"]=="color":
                color_tab.configure(fg_color=ACCENT,border_width=0,text_color="white")
                depth_tab.configure(fg_color=BG_FIELD,border_width=1,border_color=ACCENT2,text_color=ACCENT2)
            else:
                color_tab.configure(fg_color=BG_FIELD,border_width=1,border_color=ACCENT,text_color=ACCENT)
                depth_tab.configure(fg_color=ACCENT2,border_width=0,text_color="white")

        _lk=[None]
        def refresh(force_obj=False):
            total=len(matches)
            if total==0: win.destroy(); return
            i=max(0,min(state["idx"],total-1)); state["idx"]=i
            rec=matches[i]; p=rec["parts"]
            hc=bool(rec["color_path"] and os.path.isfile(rec["color_path"]))
            hd=bool(rec["depth_path"] and os.path.isfile(rec["depth_path"]))
            if state["mode"]=="color" and not hc and hd: state["mode"]="depth"
            elif state["mode"]=="depth" and not hd and hc: state["mode"]="color"
            counter_lbl.configure(text=f"{i+1} / {total}"); key_lbl.configure(text=rec["base_key"])
            seq_lbl.configure(text=f"seq {p['sequence']}")
            detail_lbl.configure(text=(f"Room {p['room']}  ·  {p['height']}  ·  "
                f"{ANGLE_LABEL.get(p['angle'],p['angle'])}  ·  {p['distance']}  ·  "
                f"{LIGHT_LABEL.get(p['lighting'],p['lighting'])}"))
            color_tab.configure(state="normal" if hc else "disabled")
            depth_tab.configure(state="normal" if hd else "disabled")
            _hl_tabs()
            path=rec["color_path"] if state["mode"]=="color" else rec["depth_path"]
            _render(path); fname_lbl.configure(text=os.path.basename(path) if path else "—")
            prev_btn.configure(state="normal" if i>0 else "disabled",text_color=TEXT_MAIN if i>0 else TEXT_DIM)
            next_btn.configure(state="normal" if i<total-1 else "disabled",text_color=TEXT_MAIN if i<total-1 else TEXT_DIM)
            if rec["base_key"]!=_lk[0] or force_obj: _lk[0]=rec["base_key"]; _rebuild_obj()

        def _on_resize(e):
            if e.widget!=canvas: return
            if state["resize_job"]: win.after_cancel(state["resize_job"])
            state["resize_job"]=win.after(150,refresh)
        canvas.bind("<Configure>",_on_resize)

        def _delete_dlg():
            i=state["idx"]; rec=matches[i]; p=rec["parts"]
            hc=bool(rec["color_path"] and os.path.isfile(rec["color_path"]))
            hd=bool(rec["depth_path"] and os.path.isfile(rec["depth_path"]))
            if not hc and not hd: messagebox.showinfo("Nothing to delete","No files on disk.",parent=win); return
            gk=group_key(p); dsq=int(p["sequence"])
            siblings=sum(1 for r in matches if group_key(r["parts"])==gk and int(r["parts"]["sequence"])>dsq)
            dlg=ctk.CTkToplevel(win); dlg.title("Delete Image"); dlg.geometry("440x370")
            dlg.resizable(False,False); dlg.configure(fg_color=BG_DARK); dlg.grab_set(); dlg.lift(); dlg.focus_force()
            ctk.CTkLabel(dlg,text="🗑  DELETE  IMAGE",font=("Courier New",12,"bold"),text_color=DANGER).pack(pady=(18,4))
            ctk.CTkLabel(dlg,text=f"Record:  {rec['base_key']}",font=("Courier New",9),text_color=TEXT_MAIN).pack()
            if siblings>0: ctk.CTkLabel(dlg,text=f"⚠  {siblings} file(s) will be renumbered.",font=("Courier New",9),text_color=WARNING).pack(pady=(4,0))
            ctk.CTkLabel(dlg,text="What to delete:",font=("Courier New",10),text_color=TEXT_DIM).pack(pady=(14,4))
            dv=tk.StringVar(value="both" if hc and hd else("color" if hc else "depth"))
            of=ctk.CTkFrame(dlg,fg_color=BG_FIELD,corner_radius=8); of.pack(fill="x",padx=20,pady=4)
            for val,txt,av in [("both","Color (.jpg) + Depth (.png)",hc and hd),
                                ("color","Color only  (.jpg)",hc),("depth","Depth only  (.png)",hd)]:
                ctk.CTkRadioButton(of,text=txt,variable=dv,value=val,font=("Courier New",11),
                    text_color=TEXT_MAIN,fg_color=DANGER,hover_color="#C94040",
                    state="normal" if av else "disabled").pack(anchor="w",padx=14,pady=6)
            rv=tk.BooleanVar(value=True)
            ctk.CTkCheckBox(dlg,text="Resequence remaining files",variable=rv,
                font=("Courier New",11),text_color=TEXT_MAIN,fg_color=ACCENT,
                hover_color=ACCENT2).pack(anchor="w",padx=20,pady=(10,4))
            ev=tk.StringVar()
            ctk.CTkLabel(dlg,textvariable=ev,font=("Courier New",9),text_color=DANGER).pack(pady=(2,0))
            def _do():
                ch=dv.get(); errs=[]
                if ch in("both","color") and rec["color_path"] and os.path.isfile(rec["color_path"]):
                    try: os.remove(rec["color_path"]); rec["color_path"]=None
                    except Exception as e: errs.append(f"Color: {e}")
                if ch in("both","depth") and rec["depth_path"] and os.path.isfile(rec["depth_path"]):
                    try: os.remove(rec["depth_path"]); rec["depth_path"]=None
                    except Exception as e: errs.append(f"Depth: {e}")
                if errs: ev.set("  |  ".join(errs)); return
                fd=(not rec["color_path"] or not os.path.isfile(rec["color_path"] or "")) and \
                   (not rec["depth_path"] or not os.path.isfile(rec["depth_path"] or ""))
                dlg.destroy()
                if rv.get() and siblings>0:
                    def _wk():
                        logs=DatasetManagerApp._resequence_on_disk(rec,matches,i,state["object_counts"])
                        self.after(0,lambda:_fin(fd,logs))
                    threading.Thread(target=_wk,daemon=True).start()
                else: _fin(fd,[])
            def _fin(fd,rl):
                ci=state["idx"]
                if fd:
                    state["object_counts"].pop(rec["base_key"],None); matches.pop(ci)
                    ch2=self._filter_tree.get_children()
                    if ci<len(ch2): self._filter_tree.delete(ch2[ci])
                    rem=self._filter_tree.get_children()
                    for j in range(ci,len(rem)):
                        ex=[t for t in self._filter_tree.item(rem[j],"tags")
                            if not t.startswith("idx:") and t not in("even","odd")]
                        self._filter_tree.item(rem[j],tags=(*ex,"even" if j%2==0 else "odd",f"idx:{j}"))
                    state["idx"]=min(ci,len(matches)-1)
                else: self._refresh_tree_row(ci)
                if rl:
                    for j in range(len(matches)): self._refresh_tree_row(j)
                self._filter_count.set(f"{len(matches)} unique record(s)")
                self._set_status(f"Deleted · {len(rl)} resequenced · {len(matches)} remaining")
                _lk[0]=None; refresh(force_obj=True)
            ctk.CTkButton(dlg,text="🗑  Confirm Delete",height=40,fg_color=DANGER,hover_color="#C94040",
                font=("Courier New",12,"bold"),text_color="white",command=_do).pack(fill="x",padx=20,pady=(12,4))
            ctk.CTkButton(dlg,text="Cancel",height=32,fg_color=BG_FIELD,border_width=1,border_color=BORDER,
                hover_color=BG_MID,font=("Courier New",11),text_color=TEXT_DIM,
                command=dlg.destroy).pack(fill="x",padx=20,pady=(0,14))

        def _save_excel():
            if not _XLSX_OK: messagebox.showerror("Missing","openpyxl required.\npip install openpyxl"); return
            all_objs=_all_objs()
            if not state["object_counts"]: messagebox.showinfo("Nothing to save","No images visited yet."); return
            auto=_excel_auto_name(matches); idir=state["save_dir"] or os.path.expanduser("~")
            path=filedialog.asksaveasfilename(initialdir=idir,initialfile=auto,
                defaultextension=".xlsx",filetypes=[("Excel file","*.xlsx")],parent=win)
            if not path: return
            state["save_dir"]=os.path.dirname(path); _write_xlsx(path,all_objs)

        def _write_xlsx(path,all_objs):
            wb=openpyxl.Workbook(); ws=wb.active; ws.title="Object Counts"
            hf=Font(name="Calibri",bold=True,color="4F8EF7",size=10)
            of2=Font(name="Calibri",bold=True,color="7C5CFC",size=10)
            df=Font(name="Calibri",bold=False,color="E8EDF5",size=10)
            cf2=Font(name="Calibri",bold=True,color="E8EDF5",size=10)
            hfill=PatternFill("solid",fgColor="1E2535"); ofill=PatternFill("solid",fgColor="252D3D")
            cen=Alignment(horizontal="center",vertical="center",wrap_text=True)
            la=Alignment(horizontal="left",vertical="center")
            thin=Side(style="thin",color="2E3A50"); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
            META=[("Date",13,la,True),("Floor",7,cen,True),("Room",7,cen,True),
                  ("Height (m)",9,cen,True),("Distance",10,cen,True),("Angle",11,cen,True),
                  ("Lighting",10,cen,True),("Resolution",12,cen,True),("RGB Format",10,cen,True),
                  ("Depth Format",11,cen,True),("Start Filename",34,la,True),
                  ("End Filename",34,la,True),("# Images",9,cen,True),("Est. Total Objects",14,cen,True)]
            OBJ=[(n,max(12,len(n)+2),cen,False) for n in all_objs]
            TRAIL=[("Object Class",18,la,True),("Notes",28,la,True)]
            ALL=META+OBJ+TRAIL; NM=len(META); NO=len(OBJ)
            for ci,(h,w,al,im) in enumerate(ALL,1):
                cell=ws.cell(row=1,column=ci,value=h)
                cell.font=of2 if not im else hf; cell.fill=ofill if not im else hfill
                cell.alignment=cen; cell.border=brd
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width=w
            ws.row_dimensions[1].height=30
            skeys=sorted(state["object_counts"].keys(),
                key=lambda k:int(next((m["parts"]["sequence"] for m in matches if m["base_key"]==k),"0")))
            for ri,key in enumerate(skeys,2):
                r2=next((m for m in matches if m["base_key"]==key),None)
                cnts=state["object_counts"].get(key,{})
                rfill=PatternFill("solid",fgColor="161B27" if ri%2==0 else "1E2535")
                if r2:
                    p2=r2["parts"]; sf2=os.path.basename(r2["color_path"]) if r2["color_path"] else ""
                    to=sum(cnts.values()); oc=", ".join(n for n in all_objs if cnts.get(n,0)>0)
                    mv=[datetime.date.today().isoformat(),p2["room"][:2],p2["room"][2:],
                        p2["height"].replace("m",""),p2["distance"].capitalize(),
                        ANGLE_LABEL.get(p2["angle"],p2["angle"]),LIGHT_LABEL.get(p2["lighting"],p2["lighting"]),
                        "1280x720","jpg","png",sf2,sf2,1,to]
                else: mv=[datetime.date.today().isoformat(),"","","","","","","","","",key,"",1,0]; oc=""
                av=mv+[cnts.get(o,0) for o in all_objs]+[oc,""]
                for ci,(val,(_,_,al,im)) in enumerate(zip(av,ALL),1):
                    cell=ws.cell(row=ri,column=ci,value=val)
                    cell.fill=rfill; cell.border=brd
                    cell.font=cf2 if not im else df; cell.alignment=cen if not im else al
                ws.row_dimensions[ri].height=18
            ws.freeze_panes="A2"
            tr=len(skeys)+2; tf2=PatternFill("solid",fgColor="0F1117")
            tfnt=Font(name="Calibri",bold=True,color="4F8EF7",size=10)
            for ci in range(1,len(ALL)+1):
                cell=ws.cell(row=tr,column=ci); cell.fill=tf2; cell.border=brd; cell.font=tfnt
                cn=ALL[ci-1][0]
                if cn=="Date": cell.value="TOTAL"; cell.alignment=la
                elif cn=="# Images": cell.value=len(skeys); cell.alignment=cen
                elif cn=="Est. Total Objects":
                    cell.value=sum(sum(state["object_counts"].get(k,{}).values()) for k in skeys); cell.alignment=cen
                elif NM<=ci-1<NM+NO:
                    on=all_objs[ci-1-NM]
                    cell.value=sum(state["object_counts"].get(k,{}).get(on,0) for k in skeys); cell.alignment=cen
            ws.row_dimensions[tr].height=20
            try:
                wb.save(path)
                messagebox.showinfo("Saved",f"Saved as:\n{os.path.basename(path)}\n\n{len(skeys)} row(s)  ·  {len(all_objs)} object column(s)",parent=win)
            except Exception as e: messagebox.showerror("Save failed",str(e),parent=win)

        def _open_excel_dlg():
            if not _XLSX_OK: messagebox.showerror("Missing","openpyxl required."); return
            idir=state["save_dir"] or os.path.expanduser("~")
            path=filedialog.askopenfilename(initialdir=idir,filetypes=[("Excel","*.xlsx"),("All","*.*")],parent=win)
            if not path: return
            stem=os.path.splitext(os.path.basename(path))[0]
            inferred=stem if re.fullmatch(r"\d{6}",stem) else None
            try: lc,on=_load_object_counts_from_xlsx(path)
            except Exception as e: messagebox.showerror("Load failed",str(e),parent=win); return
            if not lc: messagebox.showinfo("Empty file","No recognisable counts found.",parent=win); return
            for n in set(on):
                for d in state["object_counts"].values(): d.setdefault(n,0)
            overlap=0
            for bk,rc in lc.items():
                if bk not in state["object_counts"]: state["object_counts"][bk]={}
                state["object_counts"][bk].update(rc); overlap+=1
            state["save_dir"]=os.path.dirname(path); _inv_cache(); _rebuild_obj()
            jumped=False
            if inferred:
                for j,rec in enumerate(matches):
                    if rec["parts"]["room"]==inferred: state["idx"]=j; jumped=True; break
            _lk[0]=None; refresh(force_obj=True)
            info=f"Loaded {overlap} row(s) from:\n{os.path.basename(path)}\n\nObjects: {', '.join(on) or '(none)'}"
            if inferred and jumped: info+=f"\n\nJumped to room {inferred}."
            elif inferred: info+=f"\n\nRoom {inferred} not found in current results."
            messagebox.showinfo("Record loaded",info,parent=win)

        def _rename_dlg():
            i=state["idx"]; rec=matches[i]; p=rec["parts"]
            dlg=ctk.CTkToplevel(win); dlg.title("Rename File"); dlg.geometry("560x460")
            dlg.resizable(False,False); dlg.configure(fg_color=BG_DARK); dlg.grab_set(); dlg.lift(); dlg.focus_force()
            ctk.CTkLabel(dlg,text="RENAME  FILE",font=("Courier New",11,"bold"),text_color=ACCENT).pack(pady=(18,2))
            ctk.CTkLabel(dlg,text="Both color & depth files will be renamed.",font=("Courier New",9),text_color=TEXT_DIM).pack(pady=(0,12))
            form=ctk.CTkFrame(dlg,fg_color=BG_FIELD,corner_radius=8); form.pack(fill="x",padx=20,pady=4)
            vr=tk.StringVar(value=p["room"]); vh=tk.StringVar(value=p["height"])
            va=tk.StringVar(value=p["angle"]); vd=tk.StringVar(value=p["distance"])
            vl=tk.StringVar(value=p["lighting"]); vs=tk.StringVar(value=p["sequence"])
            def _row(parent,label,wfn,row):
                ctk.CTkLabel(parent,text=label,font=("Courier New",10),text_color=TEXT_DIM,anchor="e",width=130
                             ).grid(row=row,column=0,padx=(14,8),pady=8,sticky="e")
                w=wfn(parent); w.grid(row=row,column=1,padx=(0,14),pady=8,sticky="w")
            _row(form,"Floor+Room (FFRRRR)",lambda p:ctk.CTkEntry(p,textvariable=vr,width=160,font=("Courier New",11),fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=ACCENT),0)
            for r,(lbl,var,opts) in enumerate([("Height",vh,HEIGHT_OPTS),("Angle",va,ANGLE_OPTS),("Distance",vd,DIST_OPTS),("Lighting",vl,LIGHT_OPTS)],1):
                _row(form,lbl,lambda p,v=var,o=opts:ctk.CTkOptionMenu(p,variable=v,values=o,fg_color=BG_CARD,button_color=ACCENT,button_hover_color=ACCENT2,text_color=TEXT_MAIN,font=("Courier New",11),width=160),r)
            _row(form,"Sequence (4 digits)",lambda p:ctk.CTkEntry(p,textvariable=vs,width=160,font=("Courier New",11),fg_color=BG_CARD,text_color=TEXT_MAIN,border_color=ACCENT),5)
            pv=tk.StringVar()
            ctk.CTkLabel(dlg,textvariable=pv,font=("Courier New",9),text_color=SUCCESS).pack(pady=(10,0))
            def _upd(*_):
                nb={**p,"room":vr.get().strip(),"height":vh.get(),"angle":va.get(),
                    "distance":vd.get(),"lighting":vl.get(),"sequence":vs.get().strip(),
                    "is_depth":False,"ext":".jpg"}
                c=build_filename(nb); nb["is_depth"]=True; nb["ext"]=".png"; d2=build_filename(nb)
                pv.set(f"Color: {c}\nDepth: {d2}")
            for v in (vr,vh,va,vd,vl,vs): v.trace_add("write",_upd); _upd()
            ev=tk.StringVar()
            ctk.CTkLabel(dlg,textvariable=ev,font=("Courier New",9),text_color=DANGER).pack(pady=(2,0))
            def _do():
                room=vr.get().strip(); seq=vs.get().strip()
                if not re.fullmatch(r"\d{6}",room): ev.set("Room must be 6 digits."); return
                if not re.fullmatch(r"\d{4}",seq): ev.set("Sequence must be 4 digits."); return
                nb={"room":room,"height":vh.get(),"angle":va.get(),"distance":vd.get(),"lighting":vl.get(),"sequence":seq}
                errs=[]
                if rec["color_path"] and os.path.isfile(rec["color_path"]):
                    np2=os.path.join(os.path.dirname(rec["color_path"]),build_filename({**nb,"is_depth":False,"ext":".jpg"}))
                    try: os.rename(rec["color_path"],np2); rec["color_path"]=np2
                    except Exception as e: errs.append(f"Color: {e}")
                if rec["depth_path"] and os.path.isfile(rec["depth_path"]):
                    np2=os.path.join(os.path.dirname(rec["depth_path"]),build_filename({**nb,"is_depth":True,"ext":".png"}))
                    try: os.rename(rec["depth_path"],np2); rec["depth_path"]=np2
                    except Exception as e: errs.append(f"Depth: {e}")
                if errs: ev.set("  |  ".join(errs)); return
                nk=base_key(nb); ok=rec["base_key"]; rec["base_key"]=nk; rec["parts"].update(nb)
                if ok in state["object_counts"]: state["object_counts"][nk]=state["object_counts"].pop(ok)
                self._refresh_tree_row(i); dlg.destroy(); refresh(force_obj=True)
            ctk.CTkButton(dlg,text="✔  Apply Rename",height=40,fg_color=ACCENT,hover_color=ACCENT2,font=("Courier New",12,"bold"),text_color="white",command=_do).pack(fill="x",padx=20,pady=(12,4))
            ctk.CTkButton(dlg,text="Cancel",height=32,fg_color=BG_FIELD,border_width=1,border_color=BORDER,hover_color=BG_MID,font=("Courier New",11),text_color=TEXT_DIM,command=dlg.destroy).pack(fill="x",padx=20,pady=(0,14))

        def set_mode(m): state["mode"]=m; refresh()
        def go_prev():
            if state["idx"]>0: state["idx"]-=1; refresh()
        def go_next():
            if state["idx"]<len(matches)-1: state["idx"]+=1; refresh()
        color_tab.configure(command=lambda:set_mode("color"))
        depth_tab.configure(command=lambda:set_mode("depth"))
        prev_btn.configure(command=go_prev); next_btn.configure(command=go_next)
        win.bind("<Left>",lambda e:go_prev()); win.bind("<Right>",lambda e:go_next())
        win.bind("c",lambda e:set_mode("color")); win.bind("d",lambda e:set_mode("depth"))
        win.bind("<Delete>",lambda e:_delete_dlg())
        win.after(100,refresh)

    # ══════════════════════════════════════════════════════════════════════════
    #  EXPORT / COPY / SCAN
    # ══════════════════════════════════════════════════════════════════════════
    def _export_filter_list(self):
        if not self._filter_matches: messagebox.showinfo("Empty","Run a search first."); return
        path=filedialog.asksaveasfilename(defaultextension=".txt",filetypes=[("Text","*.txt"),("All","*.*")])
        if not path: return
        with open(path,"w",encoding="utf-8") as f:
            f.write("base_key\troom\theight\tangle\tdistance\tlighting\tsequence\tcolor_path\tdepth_path\n")
            for rec in self._filter_matches:
                p=rec["parts"]
                f.write(f"{rec['base_key']}\t{p['room']}\t{p['height']}\t{p['angle']}\t"
                        f"{p['distance']}\t{p['lighting']}\t{p['sequence']}\t"
                        f"{rec['color_path'] or ''}\t{rec['depth_path'] or ''}\n")
        messagebox.showinfo("Exported",f"Saved {len(self._filter_matches)} rows to\n{path}")

    def _copy_matched(self):
        if not self._filter_matches: messagebox.showinfo("Empty","Run a search first."); return
        dest=filedialog.askdirectory()
        if not dest: return
        copied=0
        for rec in self._filter_matches:
            for p in (rec["color_path"],rec["depth_path"]):
                if p and os.path.isfile(p):
                    try: shutil.copy2(p,dest); copied+=1
                    except: pass
        messagebox.showinfo("Done",f"Copied {copied} file(s) to\n{dest}")

    def _scan_all(self):
        root=self._get_root()
        if not root: return
        self._clr(self._preview_box); self._scan_count.set("Scanning…"); self._set_status("Scanning…")
        def _wk():
            lines=[]; total=bad=0
            for _,rf,fname in walk_images(root):
                p=parse_filename(fname)
                if p:
                    ft="color" if p["ext"]==".jpg" else "depth"
                    lines.append(f"  {rf:<30}  {p['room']:8}  {p['height']:7}  "
                        f"{ANGLE_LABEL.get(p['angle'],p['angle']):10}  {p['distance']:9}  "
                        f"{LIGHT_LABEL.get(p['lighting'],p['lighting']):8}  {p['sequence']:6}  {ft}")
                    total+=1
                else: lines.append(f"  [UNRECOGNISED]  {rf}/{fname}"); bad+=1
            bulk="\n".join(lines)
            def _up():
                self._preview_box.configure(state="normal")
                self._preview_box.insert("end",bulk+"\n")
                self._preview_box.configure(state="disabled")
                self._scan_count.set(f"{total} valid, {bad} unrecognised")
                self._set_status(f"Scan complete: {total} images found")
            self.after(0,_up)
        threading.Thread(target=_wk,daemon=True).start()


if __name__=="__main__":
    app=DatasetManagerApp()
    app.mainloop()
